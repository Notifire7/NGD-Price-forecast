#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ดึงข้อมูล forecast ราคาก๊าซจากไฟล์ Excel ของ Marketing -> data.json

    python3 extract.py "ForecastSP2026_....xlsx"

ปีและเส้นแบ่ง actual/forecast คำนวณเองจากวันที่ในไฟล์ (cell B1)
ไม่ต้องมาแก้โค้ดทุกเดือน
"""
import argparse, json, os, datetime as dt, sys
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string as ci

# ผังคอลัมน์ทั้งหมดอยู่ในไฟล์ groups.json ไม่ได้ฝังไว้ในโค้ด
# ทำแบบนี้เพื่อให้ไฟล์ .py ของทุกโครงการ (PTTNGD / AMATANGD) เหมือนกันทุกตัวอักษร
# ต่างกันแค่ groups.json — แก้บั๊กทีเดียวแล้วก๊อปข้ามโฟลเดอร์ได้เลย
CONFIG_DEFAULT = "groups.json"


def load_config(path):
    if not os.path.exists(path):
        sys.exit(f"[!] ไม่เจอไฟล์ผังคอลัมน์ {path}\n"
                 f"    ไฟล์นี้บอกว่าราคาแต่ละกลุ่มอยู่ชีทไหน คอลัมน์ไหน")
    cfg = json.load(open(path, encoding="utf-8"))
    for k in ("site", "actual_offset", "groups"):
        if k not in cfg:
            sys.exit(f"[!] {path} ขาดหัวข้อ '{k}'")
    for g in cfg["groups"]:
        if g.get("kind") == "cogen" and "anchor" in g:
            g["anchor"] = tuple(g["anchor"])
    return cfg

MON = ["ม.ค.", "ก.พ.", "มี.ค.", "เม.ย.", "พ.ค.", "มิ.ย.",
       "ก.ค.", "ส.ค.", "ก.ย.", "ต.ค.", "พ.ย.", "ธ.ค."]


def txt(ws, row, col):
    v = ws.cell(row, ci(col)).value
    return " ".join(str(v).split()) if v is not None else ""


def check_headers(ws, g, problems):
    """กันกรณี Marketing ย้าย/แทรกคอลัมน์แล้วเราดึงผิดช่องโดยไม่รู้ตัว"""
    def want(row, col, must, what):
        got = txt(ws, row, col)
        if must.lower() not in got.lower():
            problems.append(f"{g['sheet']} ช่อง {col}{row}: คาดว่าเป็น {what} "
                            f"(ต้องมีคำว่า '{must}') แต่เจอ '{got or '(ว่าง)'}'")

    if g["kind"] == "cogen":
        col, tag = g["anchor"]
        want(6, col, tag, f"หัวกลุ่ม {g['name']}")
        want(7, g["cols"]["Cn"], "Cn", "คอลัมน์ Cn")
        want(7, g["cols"]["Sn"], "Sn", "คอลัมน์ Sn")
        want(7, g["cols"]["price"], "Selling price", "คอลัมน์ Selling price")
    else:
        want(6, g["cols"]["Cn"], "Cn", "คอลัมน์ Cn")
        for b in g["cols"]["blocks"]:
            want(6, b["Sn"], b["label"], f"หัว {b['label']}")
            want(7, b["Sn"], "Sn", f"{b['label']} → Sn")
            want(7, b["price"], "Selling price", f"{b['label']} → Selling price")


def find_month_rows(ws, year, start_row):
    rows = {}
    for r in range(start_row, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if isinstance(v, dt.datetime) and v.year == year:
            rows[v.month] = r
    return rows


def num(ws, row, col):
    v = ws.cell(row, ci(col)).value
    try:
        return round(float(v), 2)
    except (TypeError, ValueError):
        return None


def cutoff(as_of, year, offset):
    """เดือนสุดท้ายของ 'year' ที่ถือเป็นราคาจริง (0 = ยังไม่มีเลย, 12 = จริงทั้งปี)"""
    n = (as_of.year - year) * 12 + as_of.month + offset
    return max(0, min(12, n))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx")
    ap.add_argument("--year", type=int, default=None,
                    help="ค่าเริ่มต้น = ปีของวันที่ในไฟล์")
    ap.add_argument("--out", default="data.json")
    ap.add_argument("--actual-through", metavar="YYYY-MM",
                    help="บังคับเส้นแบ่งราคาจริงเองทุก sheet (ปกติไม่ต้องใช้)")
    ap.add_argument("--no-check", action="store_true",
                    help="ข้ามการตรวจหัวคอลัมน์ (ใช้เมื่อรู้ตัวว่า Marketing เปลี่ยนผัง)")
    ap.add_argument("--groups", default=CONFIG_DEFAULT,
                    help="ไฟล์ผังคอลัมน์ (ค่าเริ่มต้น groups.json)")
    args = ap.parse_args()

    cfg = load_config(args.groups)
    GROUPS = cfg["groups"]
    ACTUAL_OFFSET = cfg["actual_offset"]
    site = cfg["site"]
    start_row = site.get("data_start_row", 8)
    date_sheet = site["date_sheet"]
    date_cell = site.get("date_cell", "B1")

    wb = load_workbook(args.xlsx, data_only=True)

    if date_sheet not in wb.sheetnames:
        sys.exit(f"[!] ไม่เจอชีท '{date_sheet}' ในไฟล์ Excel\n"
                 f"    ชีทที่มีอยู่: {', '.join(wb.sheetnames)}")

    raw = wb[date_sheet][date_cell].value
    if not isinstance(raw, dt.datetime):
        sys.exit(f"[!] อ่านวันที่จาก {date_sheet}!{date_cell} ไม่ได้ (เจอ {raw!r}) "
                 f"— ระบุ --year และ --actual-through เอง")
    as_of = raw
    year = args.year or as_of.year

    override = None
    if args.actual_through:
        y, m = (int(x) for x in args.actual_through.split("-"))
        override = cutoff(dt.datetime(y, m, 1), year, 0)

    out = {"year": year, "as_of": as_of.strftime("%Y-%m-%d"),
           "source_file": args.xlsx.split("/")[-1], "unit": "THB/MMBTU", "groups": []}

    problems, warnings = [], []

    for g in GROUPS:
        if g["sheet"] not in wb.sheetnames:
            problems.append(f"ไม่เจอชีท '{g['sheet']}' (ที่มี: {', '.join(wb.sheetnames)})")
            continue
        ws = wb[g["sheet"]]
        if not args.no_check:
            check_headers(ws, g, problems)

        rows = find_month_rows(ws, year, start_row)
        miss = [m for m in range(1, 13) if m not in rows]
        if miss:
            problems.append(f"{g['sheet']}: ไม่มีข้อมูลเดือน "
                            f"{', '.join(MON[m-1] for m in miss)} ของปี {year}")
            continue

        cut = override if override is not None else \
            cutoff(as_of, year, ACTUAL_OFFSET.get(g["sheet"], -1))

        rec = {k: g[k] for k in ("id", "name", "subtitle", "sheet", "kind")}
        rec["actual_through"] = f"{year}-{cut:02d}" if cut else None
        rec["months"] = []

        for m in range(1, 13):
            r = rows[m]
            row = {"month": m, "status": "actual" if m <= cut else "forecast",
                   "Cn": num(ws, r, g["cols"]["Cn"])}
            if g["kind"] == "cogen":
                row["Sn"] = num(ws, r, g["cols"]["Sn"])
                row["price"] = num(ws, r, g["cols"]["price"])
                if None not in (row["Cn"], row["Sn"], row["price"]):
                    d = round(row["price"] - row["Cn"] - row["Sn"], 2)
                    if abs(d) > 0.011:
                        row["epp_gap"] = d
                        warnings.append(f"{g['name']} {MON[m-1]}: Selling price ไม่เท่ากับ "
                                        f"Cn+Sn (ต่าง {d:+.2f}) — ช่อง 'ส่วนต่าง EPP' ไม่เป็น 0")
            else:
                row["blocks"] = [{"label": b["label"], "Sn": num(ws, r, b["Sn"]),
                                  "price": num(ws, r, b["price"])}
                                 for b in g["cols"]["blocks"]]
            if any(v is None for v in ([row["Cn"], row.get("Sn"), row.get("price")]
                                       if g["kind"] == "cogen"
                                       else [row["Cn"]] + [b["price"] for b in row["blocks"]])):
                warnings.append(f"{g['name']} {MON[m-1]}: มีช่องว่างในไฟล์ต้นทาง")
            rec["months"].append(row)

        out["groups"].append(rec)

    if problems:
        print("\n[!] หยุด — ผังไฟล์ไม่ตรงกับที่ตั้งค่าไว้:\n", file=sys.stderr)
        for p in problems:
            print("   • " + p, file=sys.stderr)
        print(f"\n   ถ้า Marketing เปลี่ยนผังจริง ให้แก้ไฟล์ {args.groups}\n"
              "   ถ้ามั่นใจว่าถูกแล้ว ใช้ --no-check เพื่อข้าม\n", file=sys.stderr)
        sys.exit(1)

    json.dump(out, open(args.out, "w", encoding="utf-8"), ensure_ascii=False, indent=1)

    print(f"ไฟล์ต้นทาง : {out['source_file']}")
    print(f"ข้อมูล ณ   : {as_of.strftime('%d %b %Y')}   ปีที่ดึง: {year}")
    print(f"เขียนไปที่ : {args.out}\n")
    for rec in out["groups"]:
        n = sum(1 for x in rec["months"] if x["status"] == "actual")
        edge = f"ราคาจริงถึง {MON[n-1]}" if n else "ยังไม่มีราคาจริง"
        print(f"  {rec['name']:<22} {edge:<20} ประมาณการ {12-n} เดือน")
    if warnings:
        print("\n[เตือน]")
        for w in dict.fromkeys(warnings):
            print("   • " + w)


if __name__ == "__main__":
    main()
